import os
import pandas as pd

#reading input from config.txt 
config_path = "config.txt"#path of config.txt
dff=pd.read_csv(config_path,sep=':')
data=dff[dff.columns[1]]
data=list(data)

#reading file specify its path
path='C:\\Users\\anandhu\\Desktop\\work\\count\\NEW_MR.xlsx'#path of excel file
df=pd.read_excel(path)

#filtering data base don year
df["month"]=pd.DatetimeIndex(df["TRANSACTION_DATE"]).month
df["year"]=pd.DatetimeIndex(df["TRANSACTION_DATE"]).year
df=df[df.year==int(data[6])]
df=df.replace({"month":{1:"january",2:"february",3:"march",4:"april",5:"may",6:"june",7:"july",8:"august",9:"september",10:"october",11:"november",12:"december"}})

#writing to excel
from openpyxl import Workbook

wb = Workbook()
#creating first sheet for data
ws1 = wb.active
#creating second sheet for data
ws2=wb.create_sheet("sheet2")
ws1['A1']="BASED ON REPORTED_NAME_AND_VALIDATED_LEGAL_ID"
ws1.append(["MONTHS","TOTAL","ASSIGNED","UN-ASSIGNED","ASSIGNED%","UN-ASSIGNED%","BLANK REPORTED NAME"])

#splitting month from input
mon=data[5].split(",")
a=[]
b=[]
rc1=0
rc2=0
sum1=sum2=sum3=0
#calculation of first part
for m in mon:#taking each month
    a1=[]
    a1.append(m)
    a1.append(df.loc[df.month==m,data[0]].count()) # TOTAL
    a1.append(df.loc[df.month==m,data[4]].count()) # Assigned
    a1.append(df.loc[df.month==m,data[4]].isnull().sum()) # UNASSIGNED
    if a1[1]==0:
        a1.append(0)
        a1.append(0)
    else:
        a1.append(round(a1[2]/a1[1]*100))#percentage
        a1.append(round(a1[3]/a1[1]*100))#percentage
    valm=df.loc[df.month==m,data[0]].isnull()
    recoun=0
    for val in valm:
        
        if val==True:
            recoun+=1
    a1.append(recoun)
    rc1+=recoun
    #calculating total
    sum1=sum1+a1[1]
    sum2=sum2+a1[2]
    sum3=sum3+a1[3]
    ws1.append(a1)
    
    a.append(a1)
#calculation of second part
    b1=[]
    b1.append(m)
    b1.append(df.loc[df.month==m,data[1]].sum())# TOTAL
    tf = df.loc[df[data[4]].notnull(),[data[1],"month",data[0]]]
    b1.append(tf.loc[tf.month==m,data[1]].sum())# Assigned
    b1.append(b1[1]-b1[2])# UNASSIGNED
    if b1[1]==0:
        b1.append(0)
        b1.append(0)
    else:
        b1.append(round(b1[2]/b1[1]*100))#percentage
        b1.append(round(b1[3]/b1[1]*100))#percentage
    valm=tf.loc[tf.month==m,data[0]].isnull()
    recoun2=0
    for val in valm:
        
        if val==True:
            recoun2+=1
    b1.append(recoun2)
    rc2+=recoun2
    b.append(b1)
ws1.append(["TOTAL",sum1,sum2,sum3," "," ",rc1])
ws1.append([])
ws1.append([])
ws1.append(["BASED ON REPORTED_NAME_VALIDATED_LEGAL_ID_AND_EXTENDED_SALES_PRICE","year-"+data[6]])
ws1.append(["MONTHS","TOTAL","ASSIGNED","UN-ASSIGNED","ASSIGNED%","UN-ASSIGNED%","BLANK REPORTED NAME"])

#writing second part of sheet1 to file
sum1=sum2=sum3=0
for p in b:
    ws1.append(p)
    #calucating total
    sum1=sum1+p[1]
    sum2=sum2+p[2]
    sum3=sum3+p[3]
ws1.append(["TOTAL",sum1,sum2,sum3," "," ",rc2])


ws2.append(["Count Based on REPORTED_NAME and VALIDATED_LEGAL_ID","year-"+data[6]])
ws2.append(["REGION","MONTHS","TOTAL","ASSIGNED","UN-ASSIGNED","ASSIGNED%","UN-ASSIGNED%","BLANK REPORTED NAME"])
reg=data[2].split(",")
c=[]
d=[]

#calculation based on region (1)
for r in reg:#taking each region
    ws2.append([r])
    df1=df[df[data[3]]==r]
    su1=su2=su3=0
    rc1=0
    for m in mon:#taking each month
        c1=[]
        c1.append(" ")
        c1.append(m)
        c1.append(df1.loc[df1.month==m,data[0]].count()) # TOTAL
        c1.append(df1.loc[df1.month==m,data[4]].count()) # Assigned
        c1.append(df1.loc[df1.month==m,data[4]].isnull().sum())# UNASSIGNED
        if c1[2]==0:
            c1.append(0)
            c1.append(0)
        else:
            c1.append(round(c1[3]/c1[2]*100))#percentage
            c1.append(round(c1[4]/c1[2]*100))#percentage
        valm=df1.loc[df1.month==m,data[0]].isnull()
        recoun=0
        for val in valm:
        
            if val==True:
                recoun+=1
        c1.append(recoun)
        rc1+=recoun
        #calucating total
        su1+=c1[2]
        su2+=c1[3]
        su3+=c1[4]
        ws2.append(c1)
        c.append(c1)
        
        
       
    ws2.append([" ","TOTAL",su1,su2,su3," "," ",rc1])
ws2.append(["Count Based on REPORTED_NAME _VALIDATED_LEGAL_ID and EXTENDED_SALES_PRICE"])
ws2.append(["REGION","MONTHS","TOTAL","ASSIGNED","UN-ASSIGNED","ASSIGNED%","UN-ASSIGNED%","BLANK REPORTED NAME"])
#calculation based on region (2)

for r in reg:#taking each region
    ws2.append([r])
    df1=df[df[data[3]]==r]
    su1=su2=su3=0
    rc1=0
    for m in mon:#taking each month

        d1=[]
        d1.append(" ")
        d1.append(m)
        d1.append(df1.loc[df1.month==m,data[1]].sum())# TOTAL
        tf = df1.loc[df1[data[4]].notnull(),[data[1],"month",data[0]]]
        d1.append(tf.loc[tf.month==m,data[1]].sum())# Assigned
        d1.append(d1[2]-d1[3])# UNASSIGNED
        if d1[2]==0:
            d1.append(0)
            d1.append(0)
        else:
            d1.append(round(d1[3]/d1[2]*100))#percentage
            d1.append(round(d1[4]/d1[2]*100))#percentage
        valm=tf.loc[tf.month==m,data[0]].isnull()
        recoun=0
        for val in valm:
        
            if val==True:
                recoun+=1
        rc1+=recoun
        d1.append(recoun)
        d.append(d1)
        ws2.append(d1)
        #calucating total
        su1+=d1[2]
        su2+=d1[3]
        su3+=d1[4]

    ws2.append(["","TOTAL",su1,su2,su3," "," ",rc1])
         
pn=data[8].split(",")
ws3=wb.create_sheet("sheet3")               
ws3.append(["Count Based on REPORTED_NAME and VALIDATED_LEGAL_ID","year-"+data[6]])
ws3.append(["PARTNER","REGION","MONTHS","TOTAL","ASSIGNED","UN-ASSIGNED","ASSIGNED%","UN-ASSIGNED%","BLANK REPORTED NAME"])

c=[]
d=[]

#calculation based on partner (1)
rc1=0
rc2=0
for p in pn:
    s1=s2=s3=0
    ws3.append([p])
    dataf=df[df[data[7]]==p]
    for r in reg:#taking each region
        ws3.append([" ",r])
        df1=dataf[dataf[data[3]]==r]
        su1=su2=su3=0
        for m in mon:#taking each month
            c1=[]
            c1.append(" ")
            c1.append(" ")
            c1.append(m)
            c1.append(df1.loc[df1.month==m,data[0]].count()) # TOTAL
            c1.append(df1.loc[df1.month==m,data[4]].count()) # Assigned
            c1.append(df1.loc[df1.month==m,data[4]].isnull().sum())# UNASSIGNED
            if c1[3]==0:
                c1.append(0)
                c1.append(0)
            else:
                c1.append(round(c1[4]/c1[3]*100))#percentage
                c1.append(round(c1[5]/c1[3]*100))#percentage
            valm=df1.loc[df1.month==m,data[0]].isnull()
            recoun=0
            for val in valm:
        
                if val==True:
                    recoun+=1
            c1.append(recoun)
            rc1+=recoun
            #calucating total
            su1+=c1[3]
            su2+=c1[4]
            su3+=c1[5]
            ws3.append(c1)
            c.append(c1)
        ws3.append([" "," ","REGION TOTAL",su1,su2,su3," "," ",rc1])
        rc2+=rc1
        rc1=0
        s1+=su1
        s2+=su2
        s3+=su3
    ws3.append([" "])
    ws3.append([" ","PARTNER TOTAL"," ",s1,s2,s3," "," ",rc2])
    rc2=0
    ws3.append([" "])    
       
ws3.append([" "])   
ws3.append(["Count Based on REPORTED_NAME _VALIDATED_LEGAL_ID and EXTENDED_SALES_PRICE"])
ws3.append(["PARTNER","REGION","MONTHS","TOTAL","ASSIGNED","UN-ASSIGNED","ASSIGNED%","UN-ASSIGNED%","BLANK REPORTED NAME"])

#calculation based on partner name (2)
rc1=0
rc2=0
for p in pn:
    s1=s2=s3=0
    ws3.append([p])
    dataf=df[df[data[7]]==p]
    for r in reg:#taking each region
        ws3.append(["",r])
        df1=dataf[dataf[data[3]]==r]
        su1=su2=su3=0
        for m in mon:#taking each month

            d1=[]
            d1.append(" ")
            d1.append(" ")
            d1.append(m)
            d1.append(df1.loc[df1.month==m,data[1]].sum())# TOTAL
            tf = df1.loc[df1[data[4]].notnull(),[data[1],"month",data[0]]]
            d1.append(tf.loc[tf.month==m,data[1]].sum())# Assigned
            d1.append(d1[3]-d1[4])# UNASSIGNED
            if d1[3]==0:
                d1.append(0)
                d1.append(0)
            else:
                d1.append(round(d1[4]/d1[3]*100))#percentage
                d1.append(round(d1[5]/d1[3]*100))#percentage
            valm=tf.loc[tf.month==m,data[0]].isnull()
            recoun=0
            for val in valm:
        
                if val==True:
                    recoun+=1
            d1.append(recoun)
            rc1+=recoun
            d.append(d1)
            ws3.append(d1)
            #calucating total
            su1+=d1[3]
            su2+=d1[4]
            su3+=d1[5]

        ws3.append(["","REGION TOTAL"," ",su1,su2,su3," "," ",rc1])
        rc2+=rc1
        rc1=0
        s1+=su1
        s2+=su2
        s3+=su3
    ws3.append([" "])
    ws3.append([" ","PARTNER TOTAL"," ",s1,s2,s3," "," ",rc2])
    ws3.append([" "])        
    rc2=0         
#saving excel

    
wb.save("temp.xlsx")


